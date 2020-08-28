# -*- coding:utf8 -*-
"""
Created on:
@author: BoobooWei
Email: rgweiyaping@hotmail.com
Version: V.19.03.09.0
Description:
Help:
"""

import json
import uuid
import os
import shutil
import time

# 3rd-part Modules
from openpyxl import load_workbook
from jinja2 import Template

now_date = time.strftime('%Y%m%d', time.localtime())


class OSHelper:
    def __init__(self):
        pass

    def mkdir(self, path):
        """
        在当前目录下创建子目录
        """
        try:
            os.makedirs(path)
        except Exception as e:
            # print(str(e))
            return path
        else:
            return path

    def rmdir(self, path):
        """
        递归删除目录
        """
        try:
            shutil.rmtree(path)
        except Exception as e:
            print(str(e))
            return path
        else:
            return path

    def rmfile(self, path):
        """
        删除文件
        """
        try:
            os.remove(path)
        except Exception as e:
            print(str(e))
            return path
        else:
            return path


class GetMyExcel:
    """
    读取excel数据
    """

    def __init__(self, excel):
        # read excel
        self.wb = load_workbook(excel)
        # get all sheet names
        # [u'Sheet1', u'Sheet2', u'Sheet3'}
        self.sheetnames = self.wb.sheetnames
        # print(self.sheetnames)

    def get_sheet_data(self):
        json_list = []
        for sheetname in self.sheetnames:
            lines = []
            sheet = self.wb[sheetname]
            # get row num
            # row_num = sheet.max_row
            # 此处容易出现空白， 不允许超过1000行
            row_num = 1000
            # print(row_num)
            # get column num
            column_num = sheet.max_column
            # 获取标题 title = ['a','b']
            title = list(map(lambda x: x.value, sheet['1']))
            # print(title)
            # 获取数据
            for row in range(2, row_num + 1):
                lines.append(list(map(lambda x: x.value, sheet[row])))
            for line in lines:
                if line[0] is not None:
                    json_list.append(dict(zip(title, line)))
        return json_list


class ToTopo:
    def __init__(self, data):
        self.data = data

    def get_level(self, str):
        if str == "slb":
            level = 1
        elif str == "ecs":
            level = 2
        elif str == "drds":
            level = 3
        elif str in ("rds", 'polardb', 'mongodb', 'redis'):
            level = 4
        else:
            level = 5
        return level

    def get_img(self, str):
        if str == "生产环境":
            img = "prd.png"
        elif str == "测试环境":
            img = "uat.png"
        elif str == "开发环境":
            img = "dev.png"
        elif str == "备份环境":
            img = "dr.png"
        elif str == "验证环境":
            img = "poc.png"
        else:
            img = "prd.png"
        return img

    def get_nodes(self):
        # 普通节点
        topo_json = []
        for line in self.data:
            # print(line)
            new = {
                "elementType": "node",
                "x": 300,
                "y": 500,
                "id": str(uuid.uuid4()),
                "Image": "{0}.png".format(line["产品"]),
                "text": line["实例描述"],
                "textPosition": "Bottom_Center",
                "larm": "undefined",
                "level": self.get_level(line["产品"]),
                "env": line["环境"],
                "project": line["业务"],
                "type": "normal"
            }
            # print(new)
            line.update(new)
            topo_json.append(line)

        # 环境容器和节点
        env_list = [
            {
                "elementType": "container",
                "x": 100,
                "y": 100,
                "text": "生产环境",
                "textPosition": "Middle_Center",
                "level": 2
            },
            {
                "elementType": "container",
                "x": 300,
                "y": 100,
                "text": "测试环境",
                "textPosition": "Middle_Center",
                "level": 2
            },
            {
                "elementType": "container",
                "x": 500,
                "y": 100,
                "text": "开发环境",
                "textPosition": "Middle_Center",
                "level": 2
            },
            {
                "elementType": "container",
                "x": 700,
                "y": 300,
                "text": "备份环境",
                "textPosition": "Middle_Center",
                "level": 2
            },
            {
                "elementType": "container",
                "x": 900,
                "y": 200,
                "text": "验证环境",
                "textPosition": "Middle_Center",
                "level": 2
            },
            {
                "elementType": "node",
                "x": 100,
                "y": 10,
                "id": 1000,
                "Image": "prd.png",
                "text": "生产环境",
                "textPosition": "Bottom_Center",
                "larm": "undefined",
                "level": 1,
                "env": "固定",
                "type": "env"
            },
            {
                "elementType": "node",
                "x": 300,
                "y": 10,
                "id": 3000,
                "Image": "uat.png",
                "text": "测试环境",
                "textPosition": "Bottom_Center",
                "larm": "undefined",
                "level": 1,
                "env": "固定",
                "type": "env"
            },
            {
                "elementType": "node",
                "x": 500,
                "y": 10,
                "id": 5000,
                "Image": "dev.png",
                "text": "开发环境",
                "textPosition": "Bottom_Center",
                "larm": "undefined",
                "level": 1,
                "env": "固定",
                "type": "env"
            },
            {
                "elementType": "node",
                "x": 700,
                "y": 10,
                "id": 7000,
                "Image": "dr.png",
                "text": "备份环境",
                "textPosition": "Bottom_Center",
                "larm": "undefined",
                "level": 1,
                "env": "固定",
                "type": "env"
            },
            {
                "elementType": "node",
                "x": 900,
                "y": 10,
                "id": 9000,
                "Image": "poc.png",
                "text": "验证环境",
                "textPosition": "Bottom_Center",
                "larm": "undefined",
                "level": 1,
                "env": "固定",
                "type": "env"
            }
        ]
        # 业务容器和节点
        # 统计业务
        self.projects = set(list(map(lambda x: x["project"], topo_json)))
        # print(projects)
        project_list = []
        for project in self.projects:
            project_list.extend([
                {
                    "elementType": "container",
                    "x": 100,
                    "y": 100,
                    "text": project,
                    "textPosition": "Middle_Center",
                    "level": 1
                },
            ])
        topo_json.extend(env_list + project_list)
        return topo_json

    def get_position(self, x=10, y=10, gap=0, direction='y'):
        if direction == 'y':
            x = x
            y = y + gap
        elif direction == 'x':
            x = x + gap
            y = y
        elif direction == 'xy':
            x = x + gap
            y = y + gap
        return x, y

    def auto_positon(self, topo_json):
        data = []
        data_envs = []
        data_projects = []
        data_nodes = []

        # 0. 所有的容器
        containers = list(filter(lambda x: x.get("elementType") == "container", topo_json))
        # 五个环境节点 和 N个业务节点
        # 1. 循环所有的业务
        projects = self.projects

        # 1.1 循环所有的环境
        x, y = 10, 10  # 开始画图的坐标
        for project in projects:
            # 1.1.1 画 业务节点
            # 多有的业务节点x固定为10，y暂时为10 gap=100 待最后更新y的数据
            x, y = self.get_position(x, y, 100, "y")
            project_node = {
                "elementType": "node",
                "x": x,
                "y": y,
                "id": str(uuid.uuid4()),
                "Image": "通用业务.png",
                "text": project,
                "textPosition": "Bottom_Center",
                "larm": "undefined",
                "level": 0,
                "env": "all",
                "project": project,
                "type": "normal"
            }
            print(project_node)
            data_projects.append(project_node)
            # 1.1.2 循环该业务下的所有环境 PRD UAT DEV DR PoC
            envs = list(filter(lambda x: x.get("type") == "env", topo_json))
            # print(envs)

            # 开始画该业务的第一个运行环境，坐标向x轴移动
            x_init, y_init = x, y
            # 循环所有的业务运行环境，首先打印运行环境节点
            x_max = 0
            y_max = 0
            for env in envs:
                # 从slb开始画图，固定行打印5个node
                p_nodes = []
                for p in ["slb", "ecs", "drds", "rds", "redis", "mongodb", "dts"]:
                    nodes = list(
                        filter(
                            lambda x: x.get("type") == "normal" and x.get('project') == project and x.get(
                                "env") ==
                                      env["text"] and x.get("产品") == p, topo_json))
                    p_nodes.append(nodes)
                # (slb_nodes, ecs_nodes, drds_nodes, rds_nodes, redis_nodes, mongodb_nodes, dts_nodes) = p_nodes

                # 判断该环境下是否有节点，如果没有直接continue
                if not list(filter(lambda x: x, p_nodes)):
                    break
                print("##############{}".format(p_nodes))
                # env["x"], env["y"] = x_env, y_env
                # data_envs.append(env)
                env_nodes = [{
                    "elementType": "node",
                    "x": 100,
                    "y": 10,
                    "id": str(uuid.uuid4()),
                    "Image": self.get_img(env["text"]),
                    "text": env["text"],
                    "textPosition": "Bottom_Center",
                    "larm": "undefined",
                    "level": 0,
                    "env": env["text"],
                    "project": project,
                    "type": "normal"
                }]
                print(json.dumps(env_nodes, indent=2, ensure_ascii=False))
                p_nodes.insert(0, env_nodes)

                # 开始画产品啦～ 开始的坐标为节点从环境节点之后开始向y轴移动
                x_prd, y_prd = self.get_position(x_init, y_init, 100, "x")
                for nodes in p_nodes:
                    if nodes:
                        # print(json.dumps(nodes, indent=2, ensure_ascii=False))
                        num = 0
                        start_x, start_y = x_prd, y_prd
                        p_node_x_start = x_prd
                        for node in nodes:
                            # x_prd, y_prd = x_init, y_prd + 90
                            print("###打印到第几个节点：{}".format(num))

                            if num % 5 != 0:
                                x_prd, y_prd = self.get_position(x_prd, y_prd, 200, 'x')
                            else:
                                x_prd, y_prd = self.get_position(p_node_x_start, y_prd, 90, 'y')
                            num = num + 1
                            x_max = x_max if x_prd < x_max else x_prd
                            y_max = y_max if y_prd < y_max else y_prd

                            node["x"], node["y"] = x_prd, y_prd
                            data_nodes.append(node)
                            print(
                                "{} 业务 {} 环境 {} 运行结束后的坐标 {} {}".format(project, env["text"], node.get("产品"),
                                                                       x_prd,
                                                                       y_prd))
                        # 每一个level的节点打印完成后，起点需要变化
                        if nodes[0]["level"] == 0:
                            # 如果是环境节点，则下一个开始点为当前向右斜下方运动
                            x_prd, y_prd = self.get_position(x_prd, y_prd, 60, "xy")
                        else:
                            x_prd, y_prd = self.get_position(start_x, y_prd, 40, "y")

                x_init, y_init = self.get_position(x_max, y_init, 200, "x")
                print("{} 业务 {} 环境运行结束后的坐标 {} {}".format(project, env["text"], x_prd, y_init))

            x, y = self.get_position(x, y_max, 200, "y")
            print('{} 业务运行完毕 下一次的开始坐标 {} {}'.format(project, x, y))

        data = containers + data_envs + data_projects + data_nodes
        print(json.dumps(data, indent=2, ensure_ascii=False))
        with open("online_test.json", 'w') as f:
            f.write(json.dumps(data, indent=2, ensure_ascii=False))
        return data


def startup(xlsx):
    api = GetMyExcel(xlsx)
    json_list = api.get_sheet_data()
    topo_api = ToTopo(json_list)
    topo_json = topo_api.get_nodes()
    data = topo_api.auto_positon(topo_json)

    # topo json 数据
    os_api = OSHelper()
    directory = os.getcwd()
    json_dir_name = "history"
    json_dir = os.path.join(directory, json_dir_name)
    os_api.mkdir(json_dir)
    json_file_name = '{0}-{1}.json'.format(now_date, str(uuid.uuid4()))
    path = os.path.join(json_dir, json_file_name)
    # print(path)
    # 将json文件添加到js指定的文件中
    with open("static/js/json/{json_file_name}".format(json_file_name=json_file_name), mode='w',
              encoding='utf-8') as f:
        f.write(json.dumps(data, indent=2, ensure_ascii=False))
    # 修改instances.js文件中的json文件名
    template_data = open("static/js/dynamic/excel_template.js", mode='r', encoding='utf-8').read()
    template = Template(template_data)
    j_params = {"json_file_name": json_file_name}
    with open("static/js/dynamic/excel.js".format(json_file_name=json_file_name), mode='w',
              encoding='utf-8') as f:
        f.write(template.render(**j_params))

    # 清理7天前的文件
    # 待补充


if __name__ == "__main__":

    api = GetMyExcel('history_products-20200728.xlsx')
    json_list = api.get_sheet_data()
    topo_api = ToTopo(json_list)
    topo_json = topo_api.get_nodes()
    # print(json.dumps(topo_json, indent=2, ensure_ascii=False))
    topo_api.auto_positon(topo_json)
# now_date = time.strftime('%Y%m%d', time.localtime())
#     startup('./history/excel-aaa.xlsx')
